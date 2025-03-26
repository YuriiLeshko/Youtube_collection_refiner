import os
import sys
import tempfile
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
import subprocess
from typing import List, Optional
from dataclasses import dataclass, field
from enum import Enum
from datetime import datetime, date
from functools import wraps
import logging
from contextlib import redirect_stdout, redirect_stderr
import argparse

import yt_dlp
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from moviepy.video.io.VideoFileClip import VideoFileClip


VIDEO_EXTENSIONS = [".mp4", ".mkv", ".avi", ".mov", ".flv"]
EXCEL_HEADERS = ["Video Name", "Channel", "Link", "Status", "Log"]


class Config:
    """
        Configuration class to store global settings and objects.
    """
    mode: str = "inplace"
    source_directory: Optional[os.PathLike[str]] = None
    target_directory: Optional[os.PathLike[str]] = None

    recursive: bool = True
    threads: int = 3
    retries: int = 3

    excel_path = None
    workbook = None
    sheet = None


class VideoStatus(Enum):
    UNPROCESSED = ("Unprocessed", "FFFFFF")  # White
    DOWNLOADED = ("Downloaded", "00FF00")    # Green
    MOVED = ("Moved", "FFE600")              # Yellow
    SKIPPED = ("Skipped", "C0C0C0")          # Gray
    ERROR = ("Error", "FF0000")              # Red

    def __init__(self, label, color):
        self.label = label
        self._color = color

    def __str__(self):
        return self.label

    @property
    def color(self):
        return self._color


@dataclass
class Video:
    local_path:  os.PathLike[str]
    name: str
    status: VideoStatus = VideoStatus.UNPROCESSED
    yt_id: str = 'N/A'
    yt_url: str = 'N/A'
    channel: str = 'N/A'
    duration: Optional[int] = None  # sec
    upload_date: Optional[date] = None
    log: List[str] = field(default_factory=list)
    formats: List[dict] = field(default_factory=list)


global_logger = logging.getLogger("global_logger")
global_logger.setLevel(logging.INFO)
global_logger.addHandler(logging.StreamHandler())

def retry(func):
    """
        Retry decorator to attempt a function multiple times upon failure,
        uses the retry count from a global configuration.
    """
    @wraps(func)
    def wrapper(*args, **kwargs):
        last_exception = None
        for attempt in range(Config.retries):
            try:
                result = func(*args, **kwargs)
                if result is not None:
                    return result
            except Exception as e:
                last_exception = e
                global_logger.warning(f"{func.__name__} failed on attempt {attempt + 1}: {e}")
        raise last_exception
    return wrapper


def parse_arguments():
    """
        Parse command-line arguments.
        :return: Parsed arguments object.
    """
    parser = argparse.ArgumentParser(
        description="Script to update video YouTube quality in local library and sort by channel folders."
    )
    parser.add_argument(
        '--mode',
        type=str,
        choices=['inplace', 'by-channel'],
        default='inplace',
        help="Choose how to store the processed video: 'inplace' to replace files in the same folder,"
             " 'by-channel' to create channel subfolders (default: 'inplace')."
    )
    parser.add_argument(
        '--target_dir',
        type=str,
        default=None,
        help="Target directory where processed files will be saved if mode=by-channel. "
             "Not used if mode=inplace."
    )
    parser.add_argument(
        '--source_dir',
        type=str,
        default=os.getcwd(),
        help="Source directory containing videos to process (default: current working directory)."
    )
    parser.add_argument(
        '--recursive',
        action='store_true',
        help="Enable recursive search in the source directory."
    )
    parser.add_argument(
        '--threads',
        type=int,
        default=3,
        help="Number of concurrent threads (default: 3)."
    )
    parser.add_argument(
        '--retries',
        type=int,
        default=3,
        help="Number of retries for downloading files (default: 3)."
    )
    return parser.parse_args()


def ensure_latest_package(package_name):
    """
        Ensure that the specified package is updated to the latest version using pip.
        :param package_name: Name of the package to be updated.
        :raises SystemExit: If the pip update command fails.
    """
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", package_name])
        global_logger.info(f"{package_name} successfully updated to the latest version.")
    except subprocess.CalledProcessError as e:
        global_logger.warning(f"Failed to update {package_name}: {e}")
        sys.exit(1)


def initialize_excel():
    """
        Initialize an Excel workbook to log video processing statuses.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Processing Status"
    sheet.append(EXCEL_HEADERS)

    excel_path = os.path.join(Config.source_directory, "processing_status.xlsx")
    workbook.save(excel_path)

    Config.excel_path = excel_path
    Config.workbook = workbook
    Config.sheet = sheet


def write_log_and_status(video: Video):
    """
        Write a log and status entry into the Excel file.
        :param video: Video object with details about the video to download.
    """
    excel_path = Config.excel_path
    workbook = Config.workbook
    sheet = Config.sheet

    row = [
        video.name,
        video.channel,
        video.yt_url,
        video.status.label,
        "\n".join(video.log)
    ]
    sheet.append(row)
    status_cell = sheet.cell(row=sheet.max_row, column=4)
    status_cell.fill = PatternFill(start_color=video.status.color,
                                   end_color=video.status.color,
                                   fill_type="solid")
    workbook.save(excel_path)


def format_excel_sheet():
    """
    Adjust the Excel worksheet formatting by setting column widths, enabling text wrapping for logs,
    and ensuring text visibility without dynamic row height adjustment.
    """
    sheet = Config.sheet
    columns_to_adjust = {'A': 30, 'B': 25, 'D': 15}

    for col_letter, min_width in columns_to_adjust.items():
        max_length = 0
        for cell in sheet[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max(max_length + 5, min_width)

    # Set wrap_text=True for all cells in column "Log"
    log_col_letter = 'E'
    for cell in sheet[log_col_letter]:
        cell.alignment = Alignment(wrap_text=True)
    sheet.column_dimensions[log_col_letter].width = 60

    # Set consistent row height
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 15
        for cell in row:
            if cell.column_letter != log_col_letter:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    Config.workbook.save(Config.excel_path)


def is_video_file(file_path: str) -> bool:
    """
        Check if a file is a video file based on its extension.
        :param file_path: Path to the file.
        :return: True if the file is a video, False otherwise.
    """
    return os.path.isfile(file_path) and os.path.splitext(file_path)[1].lower() in VIDEO_EXTENSIONS


def get_video_files(source_dir: os.PathLike[str], recursive: bool = True) -> list[str]:
    """
        Get a list of video files from the source directory.
        :param source_dir: Directory containing video files.
        :param recursive: Whether to search recursively in subdirectories.
        :return: List of video file paths.
    """
    if recursive:
        video_files = [
            os.path.join(root, file)
            for root, _, files in os.walk(source_dir)
            for file in files if is_video_file(file)
        ]
    else:
        video_files = [
            os.path.join(source_dir, file)
            for file in os.listdir(source_dir)
            if is_video_file(file)
        ]
    return video_files


def get_video_duration(video: Video) -> Optional[int]:
    """
        Retrieve the duration of a local video file.
        :param video: Video object containing the local path.
        :return: Duration of the video in seconds, or None if the duration cannot be retrieved.
    """
    try:
        with VideoFileClip(video.local_path) as clip:
            return int(clip.duration)
    except Exception as e:
        video.log.append(f"Unable to retrieve local video duration:{e}")
        return None


@retry
def search_youtube_video(video: Video) -> bool:
    """
        Search for a YouTube video by name and approximate duration match.
        :param video: Video object with local path and details.
        :return: True if matched, False otherwise.
    """
    search_query = f"ytsearch:{video.name}"
    video.log.append(f"Searching video...")

    with yt_dlp.YoutubeDL({'quiet': True}) as ydl:
        results = ydl.extract_info(search_query, download=False).get('entries', [])
        if not results:
            video.log.append("No results found")
            return False

        local_duration = get_video_duration(video)
        if local_duration is None:
            video.log.append("Selecting the first result.")
            matched_result = results[0]
        else:
            matched_result = None
            for res in results[:10]:
                yt_duration = res.get('duration')
                if yt_duration and abs(local_duration - yt_duration) <= 3:
                    matched_result = res
                    break
            if matched_result:
                video.log.append("Found matching duration video.")
            else:
                video.log.append("No matching duration found. Selecting the first result as fallback.")
                matched_result = results[0]

        video.yt_id = matched_result.get('id', 'N/A')
        video.yt_url = matched_result.get('webpage_url', 'N/A')
        video.channel = matched_result.get('uploader', 'N/A')
        video.duration = int(matched_result.get('duration', None))
        upload_date = matched_result.get('upload_date', None)
        if upload_date:
            video.upload_date = datetime.strptime(upload_date, '%Y%m%d').date()
        video.formats = matched_result.get('formats', [])
        video.log.append(f"Selected video from channel '{video.channel}'")
        return True


@retry
def download_youtube_format(video: Video, output_dir: str, format_id: str) -> Optional[str]:
    """
        Download a specific YouTube video format.
        :param video: Video object with details about the video to download.
        :param output_dir: Directory where the video will be saved.
        :param format_id: YouTube format id to download.
        :return: Path to the downloaded file if successful, None otherwise.
    """
    try:
        video.log.append(f"Downloading format {format_id}...")
        ydl_opts = {
            'format': format_id,
            'outtmpl': os.path.join(output_dir, '%(title)s.%(ext)s'),
            'quiet': True
        }
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info_dict = ydl.extract_info(video.yt_url, download=True)
            file_path = ydl.prepare_filename(info_dict)
            video.log.append(f"Format {format_id} downloaded successfully.")
            return file_path
    except Exception as e:
        video.log.append(f"Error downloading format {format_id}: {e}")
        return None


@retry
def merge_video_audio(video: Video, video_path: str, audio_path: str, merged_path: str) -> Optional[str]:
    """
        Merge video and audio into a single file.
        :param video: Video object with details about the video to download.
        :param video_path: Path to the video file.
        :param audio_path: Path to the audio file.
        :param merged_path: Path to save the merged file.
        :return: Path to the merged file, or None if failed.
    """
    try:
        video.log.append("Starting merge of video and audio...")
        command = [
            "ffmpeg", "-i", video_path, "-i", audio_path,
            "-c:v", "copy", "-c:a", "aac", merged_path,
        ]
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

        if result.returncode != 0:
            video.log.append(f"Ffmpeg error: {result.stderr.decode()}")
            return None

        video.log.append(f"Merged video saved to target directory")
        return merged_path
    except Exception as e:
        video.log.append(f"Error during merging: {e}")
        return None


def download_and_merge_video_audio(video: Video, target_dir: str) -> Optional[str]:
    """
        Download and merge video and audio from a YouTube video into a single file.
        :param video: Video object with details about the video to download.
        :param target_dir: Path to the folder where the video should be saved.
        :return: Path to the merged video file with 'tmp' extension, or None if the process fails.
    """
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_video_path = download_youtube_format(video, temp_dir, '136')
            temp_audio_path = download_youtube_format(video, temp_dir, '140')
            merged_path = os.path.join(target_dir, f"[TEMP]{video.name}.mp4")

            merged_result = merge_video_audio(video, temp_video_path, temp_audio_path, merged_path)
            if not merged_result:
                video.log.append("Failed to merge video and audio.")
                return None

            video.log.append("Video/audio merged and saved as [TEMP] file successfully")
            return merged_result

    except Exception as e:
        video.log.append(f"Error in download and merge process: {e}")
        return None


def video_already_exists(video: Video, target_folder: str) -> bool:
    """
        Check if a video with the given name already exists in the specified folder.
        :param video: Video object with details about the video to download.
        :param target_folder: Path to the folder where the video might exist.
        :return: True if the video exists, False otherwise.
    """
    if any(os.path.splitext(f)[0] == video.name for f in os.listdir(target_folder)):
        video.log.append(f"Video already exists in target folder. Skipping.")
        return True
    return False


@retry
def handle_local_resolution(video: Video) -> bool:
    """
        Check the resolution of a local video file.
        :param video: Video object with details about the video to download.
        :return: True if the resolution is 720p or higher, False otherwise.
    """
    try:
        with VideoFileClip(video.local_path) as vfc:
            local_resolution = vfc.size[1]  # Video height
        video.log.append(f"Local resolution is {local_resolution}.")
        if local_resolution and local_resolution >= 720:
            video.log.append(f"Local resolution is >= 720p. Moving to target folder.")
            return True
        return False
    except Exception as e:
        video.log.append(f"Error processing video resolution: {e}")
        return False


def extract_required_formats(video: Video) -> dict:
    """
        Extract required formats (22, 136, 140) from the list of available formats.
        :param video: Video object with details about the video to download.
        :return: Dictionary with required format IDs as keys and format info as values.
    """
    required_formats = {
        "22": None,    # Video + Audio, 720p
        "136": None,   # Video only, 720p
        "140": None    # Audio only
    }
    for fmt in video.formats:
        if fmt['format_id'] in required_formats:
            required_formats[fmt['format_id']] = fmt

    return required_formats


def safe_move(source_path: str, target_folder: str, file_name: str):
    """
        Safely move a file to the target directory, ensuring no overwrites.
        :param source_path: Source file path.
        :param target_folder: Target folder for the file.
        :param file_name: Name of the file.
    """
    dest_path = os.path.join(target_folder, file_name)
    counter = 1
    while os.path.exists(dest_path):
        base, ext = os.path.splitext(dest_path)
        dest_path = f"{base}_[{counter}]{ext}"
        counter += 1

    shutil.move(source_path, dest_path)


def process_video_task(file_path):
    """
       Process a single video file: check resolution, search for YouTube data, and perform appropriate actions.
       :param file_path: Path to the video file being processed.
    """
    video = Video(local_path=file_path,
                  name=os.path.splitext(os.path.basename(file_path))[0]
                  )

    global_logger.info(f"Processing video: {video.name}")

    with open(os.devnull, 'w') as o_null, redirect_stdout(o_null), redirect_stderr(o_null):
        try:
            if not search_youtube_video(video):
                video.status = VideoStatus.ERROR
                return

            if Config.mode == 'inplace':
                target_folder = os.path.dirname(video.local_path)
                if handle_local_resolution(video):
                    video.log.append(f"Local file is already 720p+, skipping download.")
                    video.status = VideoStatus.SKIPPED
                    return

            else:
                target_folder = os.path.join(Config.target_directory, video.channel)
                os.makedirs(target_folder, exist_ok=True)

                if video_already_exists(video, target_folder):
                    os.remove(video.local_path)
                    video.log.append(f"Old file deleted successfully.")
                    video.status = VideoStatus.SKIPPED
                    return

                if handle_local_resolution(video):
                    shutil.move(video.local_path, target_folder)
                    video.status = VideoStatus.MOVED
                    return

            formats = extract_required_formats(video)
            video.log.append(f"Available formats: {', '.join(formats)}")

            if formats["136"] and formats["140"]:
                video.log.append("Merging formats 136 and 140...")
                merged_path = download_and_merge_video_audio(video, target_folder)
                if merged_path:
                    if Config.mode == 'inplace':
                        os.replace(merged_path, video.local_path)
                        video.log.append(f"File replaced in-place successfully.")
                    else:
                        os.rename(merged_path, os.path.join(target_folder, f"{video.name}.mp4"))
                        video.log.append(f"Merged [Temp] file renamed successfully.")
                    video.status = VideoStatus.DOWNLOADED
                else:
                    video.status = VideoStatus.ERROR
            else:
                video.log.append("Required formats are not available. Using existing file.")
                if Config.mode == 'by-channel':
                    shutil.move(video.local_path, target_folder)
                    video.log.append("Existing file moved to channel folder.")
                    video.status = VideoStatus.MOVED
                else:
                    video.log.append("Mode=inplace; leaving file as is.")
                    video.status = VideoStatus.SKIPPED

        except Exception as e:
            video.log.append(f"Error processing video: {e}")
            video.status = VideoStatus.ERROR

        finally:
            write_log_and_status(video)


def process_videos():
    """
        Process all video files in the source directory, utilizing multithreading for efficiency.
    """

    video_files = get_video_files(Config.source_directory, Config.recursive)

    global_logger.info(f"Found {len(video_files)} videos for processing.")

    with ThreadPoolExecutor(Config.threads) as executor:
        futures = {
            executor.submit(process_video_task, file_path):
                file_path for file_path in video_files
        }

        for future in as_completed(futures):
            video_path = futures[future]
            try:
                future.result()
                global_logger.info(f"Processing completed for: {video_path}")
            except Exception as e:
                global_logger.warning(f"Error processing {video_path}: {e}")


if __name__ == '__main__':
    ensure_latest_package("yt-dlp")
    ensure_latest_package("ffmpeg")

    params = parse_arguments()

    Config.mode = params.mode
    Config.source_directory = params.source_dir
    Config.recursive = params.recursive
    Config.threads = params.threads
    Config.retries = params.retries

    if Config.mode == 'inplace':
        Config.target_directory = Config.source_directory
    else:
        if not params.target_dir:
            sys.exit("Error: --target_dir is required if --mode=by-channel.")
        Config.target_directory = params.target_dir

    initialize_excel()
    process_videos()
    format_excel_sheet()

    global_logger.info(f"\nProcessing completed. \nReport saved at {Config.excel_path}")
